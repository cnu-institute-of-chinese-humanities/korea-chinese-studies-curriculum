# -*- coding: utf-8 -*-
"""
한국 대학의 중국학 유관 학과 교과과정 데이터셋 — 전처리 스크립트

data/raw/*.xlsx (원본, 수정하지 않음)
  -> data/processed/csv/*.csv       엑셀·SPSS·R 사용자용
  -> data/processed/json/*.json     다중값을 배열로 표현
  -> data/processed/graph.json      통합 지식그래프 {nodes, edges}
  -> data/processed/campus.geojson  캠퍼스 위치 (지도용)
  -> data/processed/_columns.json   열 정의 (코드북 작성 근거)

실행:  python scripts/preprocess.py
"""

import csv
import glob
import json
import os
import re
import sys
from collections import OrderedDict, defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다:  pip install openpyxl")


# ---------------------------------------------------------------
# 설정
# ---------------------------------------------------------------

# 교수 이름 익명화 방식. 연구팀 결정에 따라 이 값만 바꾸고 다시 실행하면 된다.
#
#   "asis"       원본 그대로                                   <- 현재 설정
#                2026-09-03 결정: 원본 시트 자체에서 교원 성명 435건과 Edge의
#                SourceName 1,011건을 성씨 없이 "○○○"으로 통일했다. 따라서 원본을
#                그대로 내보내는 것이 맞다. 아래 assert_anonymized()가 성씨가 남은
#                값이 하나라도 있으면 실행을 중단한다.
#   "pseudonym"  이름을 식별자로 대체 (Prof001)
#   "drop"       Professor 시트와 관련 Edge를 산출물에서 제외
#
ANON_MODE = "asis"

# 완전히 가려진 이름으로 인정하는 형태. 이 밖의 값이 있으면 공개용으로 부적절하다.
MASKED_NAME = re.compile(r"^[○◯]+$")

# 머리글이 없는 열에 부여하는 이름. {시트: {0-기준 열 번호: (영문 키, 한글명)}}
#
# Class_Course의 V·W열은 원본 시트에 머리글이 없다. 2026-09-03 결정에 따라 공개
# 데이터셋에 포함하되, 이름은 여기서 부여한다. 머리글이 없는 다른 열은 건너뛴다.
UNLABELED_COLUMNS = {
    "Class_Course": {
        21: ("note", "수집 메모"),
        22: ("previousDescription", "이전 교과목 해설"),
    },
}

# 문자로 적힌 결측. {열 키: {결측으로 볼 값}}
# previousDescription의 "없음" 37건은 "해설이 없다"는 뜻이므로 빈 칸으로 통일한다.
LITERAL_MISSING = {
    "previousDescription": {"없음"},
}

# 산출물에서 제외하는 값. {열 키: {값}}
# V1410의 메모는 시트 오른쪽 열(W)을 가리키는 시트 구조 주석이다. W열에 이름이
# 붙으면 뜻이 없어지므로 뺀다.
DROP_VALUES = {
    "note": {"우측: 이전 교과목 해설 정보"},
}

# 시트 이름 -> 출력 파일 이름
SHEETS = OrderedDict([
    ("Class_University", "university"),
    ("Class_Campus",     "campus"),
    ("Class_College",    "college"),
    ("Class_Faculty",    "faculty"),
    ("Class_Department", "department"),
    ("Class_Major",      "major"),
    ("Class_Track",      "track"),
    ("Class_EduGoal",    "edugoal"),
    ("Class_Course",     "course"),
    ("Class_Professor",  "professor"),
    ("Edge",             "edge"),
])

# 정수로 변환할 열
INT_KEYS = {"no", "aYear", "eYear", "oYear", "curriumYear", "grade", "semester"}
# 숫자로 변환할 열. 정수면 정수로, 아니면 실수로 둔다.
# credit에는 0.25·0.5 학점이 있으므로(30건) 정수로 강제하면 0이 되어 버린다.
NUMBER_KEYS = {"credit"}
# 실수로 변환할 열
FLOAT_KEYS = {"latitude", "longitude"}
# 세미콜론 구분 다중값 -> JSON에서 배열로
MULTI_KEYS = {"wideField", "detailedField", "categoryA", "categoryB",
              "referenceURL", "descriptionSourceURL"}

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(ROOT, "data", "raw")
OUT_DIR = os.path.join(ROOT, "data", "processed")


# ---------------------------------------------------------------
# 값 정리
# ---------------------------------------------------------------

def strip_tracking(url):
    """URL에서 Google Analytics 추적 파라미터를 제거한다.

    원본에는 자료를 수집한 사람의 브라우저 추적 ID(_ga, _gid 등)가 붙어 있다.
    개인을 식별하는 정보는 아니지만 공개 자료에 남길 이유가 없다.
    """
    if not url:
        return url
    url = re.sub(r"[?&]_g[a-z]+=[^&\s;]*", "", url)
    return re.sub(r"[?&]+$", "", url)


def clean(value):
    """셀 값을 문자열로 정규화한다. 빈 값은 None."""
    if value is None:
        return None
    # 엑셀은 정수도 실수(1.0)로 읽히므로, 정수값이면 소수점을 떼고 문자열로 만든다.
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    if s == "" or s.lower() in ("nan", "none"):
        return None
    s = re.sub(r"\s+", " ", s.replace(" ", " "))
    if "http" in s:
        s = strip_tracking(s)
    return s or None


def split_multi(s):
    """'중국문화학;중국현대문학' -> ['중국문화학', '중국현대문학']"""
    if s is None:
        return None
    parts = [p.strip() for p in s.split(";")]
    parts = [p for p in parts if p]
    return parts or None


def coerce(key, s):
    """열 이름에 따라 자료형을 맞춘다. 변환 실패 시 원래 문자열을 유지한다."""
    if s is None:
        return None
    if key in INT_KEYS:
        try:
            return int(float(s))
        except ValueError:
            return s
    if key in NUMBER_KEYS:
        try:
            f = float(s)
            return int(f) if f.is_integer() else f
        except ValueError:
            return s
    if key in FLOAT_KEYS:
        try:
            return float(s)
        except ValueError:
            return s
    return s


# ---------------------------------------------------------------
# 익명화
# ---------------------------------------------------------------

def anonymize_professor(records):
    """ANON_MODE에 따라 교수 이름 열을 처리한다."""
    if ANON_MODE == "asis":
        return records
    for r in records:
        if ANON_MODE == "pseudonym":
            r["name"] = r.get("id")
    return records


def assert_anonymized(tables):
    """교원 이름이 어디에도 남아 있지 않은지 확인하고, 남아 있으면 중단한다.

    Professor의 name과 Edge의 SourceName·TargetName(교원 쪽)을 본다.
    "○○○"처럼 전부 가려진 값이나 식별자(Prof001)만 허용한다.
    """
    bad = []
    if "professor" in tables:
        for r in tables["professor"][1]:
            v = r.get("name")
            if v and not MASKED_NAME.match(v) and v != r.get("id"):
                bad.append(("professor.name", r.get("id"), v))
    if "edge" in tables:
        for r in tables["edge"][1]:
            for cls, nm in (("SourceClass", "SourceName"), ("TargetClass", "TargetName")):
                if r.get(cls) == "Professor":
                    v = r.get(nm)
                    if v and not MASKED_NAME.match(v) and not v.startswith("Prof"):
                        bad.append(("edge." + nm, r.get("no"), v))
    if bad:
        print("\n  중단: 교원 이름이 가려지지 않은 값 %d건" % len(bad))
        for where, key, v in bad[:10]:
            print("      %s  %s  %s" % (where, key, v))
        sys.exit(1)
    print("  익명화 확인 : 교원 이름 잔존 0건")


# ---------------------------------------------------------------
# 읽기
# ---------------------------------------------------------------

def read_sheet(ws):
    """시트를 (열정의, 레코드목록)으로 읽는다.

    헤더 셀은 영문 키와 한글 설명이 줄바꿈으로 붙어 있다. 예: no + 연번
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    unlabeled = UNLABELED_COLUMNS.get(ws.title, {})
    columns = []
    for idx, cell in enumerate(rows[0]):
        if cell is None or not str(cell).strip():
            if idx in unlabeled:
                key, label = unlabeled[idx]
                columns.append({"index": idx, "key": key, "label": label})
            continue
        parts = [p.strip() for p in str(cell).split("\n") if p.strip()]
        columns.append({
            "index": idx,
            "key": parts[0],
            "label": " ".join(parts[1:]) if len(parts) > 1 else "",
        })

    records = []
    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rec = OrderedDict()
        for col in columns:
            i, k = col["index"], col["key"]
            v = clean(row[i]) if i < len(row) else None
            if v is not None and v in LITERAL_MISSING.get(k, ()):
                v = None
            if v is not None and v in DROP_VALUES.get(k, ()):
                v = None
            rec[k] = v
        records.append(rec)
    return columns, records


# ---------------------------------------------------------------
# 쓰기
# ---------------------------------------------------------------

def write_csv(path, columns, records):
    """CSV는 원본에 가깝게. 다중값은 세미콜론으로 유지하되 구분자를 통일한다."""
    keys = [c["key"] for c in columns]
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(keys)
        for rec in records:
            out = []
            for k in keys:
                v = rec.get(k)
                if v is None:
                    out.append("")
                elif k in MULTI_KEYS:
                    parts = split_multi(v)
                    out.append(";".join(parts) if parts else "")
                else:
                    out.append(v)
            w.writerow(out)


def write_json(path, columns, records):
    """JSON은 다중값을 배열로, 결측을 null로 표현한다."""
    out = []
    for rec in records:
        obj = OrderedDict()
        for c in columns:
            k = c["key"]
            v = rec.get(k)
            obj[k] = split_multi(v) if k in MULTI_KEYS else coerce(k, v)
        out.append(obj)
    dump(path, out)


def dump(path, obj):
    """한글을 그대로, 들여쓰기 2칸으로 저장한다.

    ensure_ascii=False 가 없으면 유니코드 이스케이프로 저장되어 사람이 읽을 수 없다.
    한 줄로 압축하면 git이 변경 이력을 행 단위로 보여주지 못한다.
    """
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
        f.write("\n")


# ---------------------------------------------------------------
# 메인
# ---------------------------------------------------------------

def main():
    src = sorted(glob.glob(os.path.join(RAW_DIR, "*.xlsx")))
    if not src:
        sys.exit("data/raw/ 에 xlsx 파일이 없습니다.")
    src = src[0]

    print("원본        :", os.path.basename(src))
    print("익명화 방식 :", ANON_MODE)
    if ANON_MODE == "asis":
        print("  (원본 시트에서 이미 익명화됨 — 아래에서 잔존 여부를 검사한다)")
    print()

    wb = openpyxl.load_workbook(src, data_only=True)

    for d in ("csv", "json"):
        os.makedirs(os.path.join(OUT_DIR, d), exist_ok=True)

    tables = OrderedDict()

    for sheet, name in SHEETS.items():
        if sheet not in wb.sheetnames:
            print("  건너뜀 (시트 없음):", sheet)
            continue
        columns, records = read_sheet(wb[sheet])

        if sheet == "Class_Professor":
            if ANON_MODE == "drop":
                print("  %-14s 제외됨 (ANON_MODE=drop)" % name)
                continue
            records = anonymize_professor(records)

        if sheet == "Edge" and ANON_MODE == "pseudonym":
            for r in records:
                if r.get("SourceClass") == "Professor":
                    r["SourceName"] = r.get("SourceID")
                if r.get("TargetClass") == "Professor":
                    r["TargetName"] = r.get("TargetID")

        if sheet == "Edge" and ANON_MODE == "drop":
            before = len(records)
            records = [r for r in records
                       if r.get("SourceClass") != "Professor"
                       and r.get("TargetClass") != "Professor"]
            print("  %-14s Professor 관계 %d건 제외" % (name, before - len(records)))

        tables[name] = (columns, records)
        write_csv(os.path.join(OUT_DIR, "csv", name + ".csv"), columns, records)
        write_json(os.path.join(OUT_DIR, "json", name + ".json"), columns, records)
        print("  %-14s %6d행 x %2d열" % (name, len(records), len(columns)))

    print()
    assert_anonymized(tables)
    build_graph(tables)
    build_geojson(tables)
    write_columns(tables)
    check_integrity(tables)

    wb.close()
    print("\n출력 : data/processed/")


def build_graph(tables):
    """노드/엣지 통합 그래프. Gephi, Cytoscape, networkx, D3에서 바로 읽힌다."""
    nodes, edges = [], []
    for name, (columns, records) in tables.items():
        if name == "edge":
            continue
        for rec in records:
            attrs = OrderedDict()
            for c in columns:
                k = c["key"]
                if k in ("no", "class", "id", "name"):
                    continue
                v = rec.get(k)
                attrs[k] = split_multi(v) if k in MULTI_KEYS else coerce(k, v)
            nodes.append(OrderedDict([
                ("id", rec.get("id")),
                ("class", rec.get("class")),
                ("label", rec.get("name")),
                ("attributes", attrs),
            ]))

    if "edge" in tables:
        for rec in tables["edge"][1]:
            edges.append(OrderedDict([
                ("source", rec.get("SourceID")),
                ("target", rec.get("TargetID")),
                ("relationship", rec.get("Relationship")),
                ("sourceClass", rec.get("SourceClass")),
                ("targetClass", rec.get("TargetClass")),
                ("description", rec.get("description")),
            ]))

    dump(os.path.join(OUT_DIR, "graph.json"), OrderedDict([
        ("directed", True),
        ("anonymization", ANON_MODE),
        ("nodeCount", len(nodes)),
        ("edgeCount", len(edges)),
        ("nodes", nodes),
        ("edges", edges),
    ]))
    print("\n  graph.json      노드 %d개, 엣지 %d개" % (len(nodes), len(edges)))


def build_geojson(tables):
    """캠퍼스 위경도 -> GeoJSON. QGIS·지도 서비스에 그대로 올릴 수 있다."""
    if "campus" not in tables:
        return
    feats = []
    for rec in tables["campus"][1]:
        lat = coerce("latitude", rec.get("latitude"))
        lon = coerce("longitude", rec.get("longitude"))
        if not isinstance(lat, float) or not isinstance(lon, float):
            continue
        props = OrderedDict((k, v) for k, v in rec.items()
                            if k not in ("latitude", "longitude"))
        feats.append(OrderedDict([
            ("type", "Feature"),
            ("geometry", OrderedDict([("type", "Point"), ("coordinates", [lon, lat])])),
            ("properties", props),
        ]))
    dump(os.path.join(OUT_DIR, "campus.geojson"),
         OrderedDict([("type", "FeatureCollection"), ("features", feats)]))
    print("  campus.geojson  지점 %d개" % len(feats))


def write_columns(tables):
    """열 정의와 결측률. docs/codebook.md 작성의 근거 자료."""
    out = OrderedDict()
    for name, (columns, records) in tables.items():
        items = []
        for c in columns:
            k = c["key"]
            filled = sum(1 for r in records if r.get(k) is not None)
            dtype = ("integer" if k in INT_KEYS else
                     "number" if k in FLOAT_KEYS or k in NUMBER_KEYS else
                     "array<string>" if k in MULTI_KEYS else "string")
            items.append(OrderedDict([
                ("key", k),
                ("label", c["label"]),
                ("type", dtype),
                ("filled", filled),
                ("total", len(records)),
                ("missingRate", round((len(records) - filled) / len(records), 3) if records else 0),
            ]))
        out[name] = OrderedDict([("rows", len(records)), ("columns", items)])
    dump(os.path.join(OUT_DIR, "_columns.json"), out)
    print("  _columns.json   열 정의 %d개 시트" % len(out))


def check_integrity(tables):
    """Edge가 가리키는 노드가 실제로 존재하는지 확인한다."""
    if "edge" not in tables:
        return
    ids = set()
    for name, (_, records) in tables.items():
        if name == "edge":
            continue
        ids.update(r.get("id") for r in records if r.get("id"))

    dangling = defaultdict(int)
    for rec in tables["edge"][1]:
        for side in ("SourceID", "TargetID"):
            v = rec.get(side)
            if v and v not in ids:
                dangling[v] += 1

    print("\n  참조 무결성 : ", end="")
    if dangling:
        print("존재하지 않는 id를 가리키는 엣지 %d종 %d건" % (len(dangling), sum(dangling.values())))
        for k, v in sorted(dangling.items(), key=lambda x: -x[1])[:10]:
            print("      %s (%d건)" % (k, v))
    else:
        print("이상 없음")


if __name__ == "__main__":
    main()
